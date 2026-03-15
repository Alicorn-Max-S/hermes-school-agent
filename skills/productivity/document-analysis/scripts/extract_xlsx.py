#!/usr/bin/env python3
"""Extract data from Excel files using openpyxl (FREE, local).

Usage:
    python3 extract_xlsx.py spreadsheet.xlsx
    python3 extract_xlsx.py spreadsheet.xlsx --sheet "Sheet1"
    python3 extract_xlsx.py spreadsheet.xlsx --format csv
    python3 extract_xlsx.py spreadsheet.xlsx --format summary
"""

import json
import sys
from pathlib import Path


def extract(input_path, sheet_name=None, fmt="summary"):
    """Extract data from an Excel file."""
    path = Path(input_path)
    if not path.exists():
        return {"success": False, "error": f"File not found: {input_path}"}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "success": False,
            "error": "openpyxl not installed. Run: pip install openpyxl",
        }

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "error": f"Failed to open Excel file: {e}"}

    sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames
    result_sheets = []

    for sname in sheets_to_process:
        if sname not in wb.sheetnames:
            result_sheets.append({"sheet": sname, "error": f"Sheet '{sname}' not found"})
            continue

        ws = wb[sname]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            result_sheets.append({"sheet": sname, "rows": 0, "columns": 0, "content": ""})
            continue

        if fmt == "csv":
            import csv
            import io
            output = io.StringIO()
            writer = csv.writer(output)
            for row in rows:
                writer.writerow(row)
            content = output.getvalue()
        elif fmt == "summary":
            headers = rows[0] if rows else []
            num_rows = len(rows) - 1  # exclude header
            num_cols = len(headers)

            # Sample data types from first few data rows
            preview_rows = rows[1:6]  # first 5 data rows
            tail_rows = rows[-3:] if len(rows) > 6 else []

            lines = [f"Sheet: {sname}", f"Columns: {num_cols}, Data Rows: {num_rows}", ""]

            # Header row
            if headers:
                lines.append("Headers: " + " | ".join(headers))
                lines.append("")

            # Preview
            if preview_rows:
                lines.append("First rows:")
                for row in preview_rows:
                    lines.append("  " + " | ".join(row))

            if tail_rows and len(rows) > 6:
                lines.append(f"\n  ... ({num_rows - 5} more rows) ...\n")
                lines.append("Last rows:")
                for row in tail_rows:
                    lines.append("  " + " | ".join(row))

            content = "\n".join(lines)
        else:
            # Default: tab-separated
            content = "\n".join("\t".join(row) for row in rows)

        result_sheets.append({
            "sheet": sname,
            "rows": len(rows),
            "columns": len(rows[0]) if rows else 0,
            "content": content,
        })

    wb.close()

    return {
        "success": True,
        "sheets": result_sheets,
        "sheet_names": wb.sheetnames,
        "format": fmt,
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: extract_xlsx.py <file> [--sheet name] [--format csv|summary]"}))
        sys.exit(1)

    sheet = None
    fmt = "summary"
    args = sys.argv[2:]
    i = 0
    while i < len(args):
        if args[i] == "--sheet" and i + 1 < len(args):
            sheet = args[i + 1]
            i += 2
        elif args[i] == "--format" and i + 1 < len(args):
            fmt = args[i + 1]
            i += 2
        else:
            i += 1

    result = extract(sys.argv[1], sheet, fmt)
    print(json.dumps(result, indent=2, ensure_ascii=False))
